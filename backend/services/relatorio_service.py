from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, Dict
from datetime import date
import io

from models.checklist import Checklist
from models.veiculo import Veiculo
from models.usuario import Colaborador


def _aplicar_filtros(query, filtros: Dict):
    if filtros.get("data_inicio"):
        query = query.filter(Checklist.data_checklist >= filtros["data_inicio"])
    if filtros.get("data_fim"):
        query = query.filter(Checklist.data_checklist <= filtros["data_fim"])
    if filtros.get("veiculo_id"):
        query = query.filter(Checklist.veiculo_id == filtros["veiculo_id"])
    if filtros.get("colaborador_id"):
        query = query.filter(Checklist.colaborador_id == filtros["colaborador_id"])
    if filtros.get("status"):
        query = query.filter(Checklist.status == filtros["status"])
    return query


def gerar_relatorio_excel(db: Session, filtros: Dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlalchemy.orm import joinedload

    query = db.query(Checklist).join(Veiculo).join(Colaborador, Checklist.colaborador_id == Colaborador.id).options(
        joinedload(Checklist.veiculo).joinedload(Veiculo.responsavel_manutencao),
        joinedload(Checklist.colaborador),
        joinedload(Checklist.validacao),
    )
    query = _aplicar_filtros(query, filtros)
    checklists = query.order_by(Checklist.data_checklist.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklists"

    cabecalho = [
        "ID", "Data", "Placa", "Modelo", "Nº Frota",
        "Colaborador", "KM Atual", "Status",
        "Doc. em dia", "Equipamentos", "Avarias", "Apto para uso", "Pneus OK",
        "Resp. Manutenção",
        "Observação", "Validado em", "Justificativa"
    ]

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_cores = {"aprovado": "C6EFCE", "reprovado": "FFC7CE", "pendente": "FFEB9C"}

    for row, c in enumerate(checklists, 2):
        respostas = c.respostas or {}
        validacao = c.validacao

        linha = [
            c.id,
            c.data_checklist.isoformat() if c.data_checklist else "",
            c.veiculo.placa if c.veiculo else "",
            c.veiculo.modelo if c.veiculo else "",
            c.veiculo.numero_frota if c.veiculo else "",
            c.colaborador.nome if c.colaborador else "",
            c.km_atual,
            c.status,
            "Sim" if respostas.get("documentacao_em_dia") else "Não",
            "Sim" if respostas.get("equipamentos_obrigatorios") else "Não",
            "Sim" if respostas.get("avarias_visiveis") else "Não",
            "Sim" if respostas.get("apto_para_uso") else "Não",
            "Sim" if respostas.get("pneus_condicao_adequada") else "Não",
            c.veiculo.responsavel_manutencao.nome if c.veiculo and c.veiculo.responsavel_manutencao else "",
            c.observacao or "",
            validacao.validado_em.strftime("%d/%m/%Y %H:%M") if validacao else "",
            validacao.justificativa if validacao else ""
        ]

        cor = status_cores.get(c.status, "FFFFFF")
        fill = PatternFill("solid", start_color=cor)

        for col, valor in enumerate(linha, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            if col == 8:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_relatorio_pdf(db: Session, filtros: Dict) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from sqlalchemy.orm import joinedload

    query = db.query(Checklist).join(Veiculo).join(Colaborador, Checklist.colaborador_id == Colaborador.id).options(
        joinedload(Checklist.veiculo).joinedload(Veiculo.responsavel_manutencao),
        joinedload(Checklist.colaborador),
    )
    query = _aplicar_filtros(query, filtros)
    checklists = query.order_by(Checklist.data_checklist.desc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Relatório de Checklist de Frota", styles["Title"]))
    elements.append(Spacer(1, 0.5*cm))

    total = len(checklists)
    aprovados = sum(1 for c in checklists if c.status == "aprovado")
    reprovados = sum(1 for c in checklists if c.status == "reprovado")
    pendentes = sum(1 for c in checklists if c.status == "pendente")

    resumo_data = [
        ["Total", "Aprovados", "Reprovados", "Pendentes"],
        [str(total), str(aprovados), str(reprovados), str(pendentes)]
    ]
    resumo_table = Table(resumo_data, colWidths=[4*cm]*4)
    resumo_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.lightgrey, colors.white])
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.5*cm))

    cabecalho = ["ID", "Data", "Placa", "Colaborador", "KM", "Status", "Avarias", "Resp. Manutenção", "Observação"]
    dados = [cabecalho]
    for c in checklists:
        respostas = c.respostas or {}
        dados.append([
            str(c.id),
            c.data_checklist.strftime("%d/%m/%Y") if c.data_checklist else "",
            c.veiculo.placa if c.veiculo else "",
            c.colaborador.nome if c.colaborador else "",
            str(c.km_atual),
            c.status.upper(),
            "SIM" if respostas.get("avarias_visiveis") else "NÃO",
            c.veiculo.responsavel_manutencao.nome if c.veiculo and c.veiculo.responsavel_manutencao else "",
            (c.observacao or "")[:40]
        ])

    tabela = Table(dados, colWidths=[1.2*cm, 2.2*cm, 2.5*cm, 4*cm, 2*cm, 2.5*cm, 2*cm, 3*cm, 5*cm])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")])
    ]))
    elements.append(tabela)

    doc.build(elements)
    return buffer.getvalue()


def montar_resumo_semanal(db: Session) -> Dict:
    from datetime import timedelta
    hoje = date.today()
    inicio = hoje - timedelta(days=7)

    total = db.query(func.count(Checklist.id)).filter(Checklist.data_checklist >= inicio).scalar()
    aprovados = db.query(func.count(Checklist.id)).filter(Checklist.data_checklist >= inicio, Checklist.status == "aprovado").scalar()
    reprovados = db.query(func.count(Checklist.id)).filter(Checklist.data_checklist >= inicio, Checklist.status == "reprovado").scalar()
    pendentes = db.query(func.count(Checklist.id)).filter(Checklist.data_checklist >= inicio, Checklist.status == "pendente").scalar()

    veiculos_ativos = db.query(func.count(Veiculo.id)).filter(Veiculo.situacao == "Ativo").scalar()
    com_checklist = db.query(Checklist.veiculo_id).filter(Checklist.data_checklist == hoje).distinct().count()

    return {
        "total": total,
        "aprovados": aprovados,
        "reprovados": reprovados,
        "pendentes": pendentes,
        "sem_checklist": veiculos_ativos - com_checklist
    }
